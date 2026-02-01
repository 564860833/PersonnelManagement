import pandas as pd
import logging
import re
import os  # 添加os模块导入
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import load_workbook
import xlrd

from database import Database

logger = logging.getLogger('ExcelImport')


def clean_column_name(name: str) -> str:
    """清理Excel列名，处理空格和换行符，保留特殊符号"""
    name = re.sub(r'[\s\u3000\n]+', '', name)
    cleaned = re.sub(r'[()（）]', '', name)  # 移除括号，但保留斜杠
    return cleaned


def convert_excel_date(val: Any) -> str:
    """
    严格保持原始Excel格式，保留所有换行符和特殊字符

    参数:
        val: Excel单元格中的原始值

    返回:
        字符串格式的原始值，保留所有换行符和特殊字符
    """
    # 处理空值
    if pd.isna(val):
        return ''

    # 对于字符串类型，直接返回（保留所有换行符）
    if isinstance(val, str):
        return val

    # 对于日期类型，转换为YYYY.MM格式
    if isinstance(val, datetime):
        return val.strftime("%Y.%m")

    # 对于数值类型（整数、浮点数），转换为字符串并保留小数部分
    if isinstance(val, (int, float)):
        # 保留原始数值格式（不截断小数点后的0）
        return str(val)

    # 其他类型转为字符串
    return str(val)


def import_specific_table(file_path: str, db: Database, table_name: str) -> (bool, str):
    """将Excel文件导入到指定数据库表，支持合并单元格处理"""
    # 表名到中文的映射
    table_name_mapping = {
        'base_info': '人员基本信息',
        'rewards': '人员奖惩信息',
        'family': '人员家庭成员信息',
        'resume': '人员简历信息'
    }

    if table_name not in table_name_mapping:
        return False, f"无效的表名: {table_name}"

    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            return False, f"文件不存在: {file_path}"

        # 检查文件格式
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext not in ['.xlsx', '.xls']:
            return False, f"不支持的文件格式: {file_ext}，请使用.xlsx或.xls格式"

        # 尝试打开文件
        try:
            with open(file_path, 'rb') as test_file:
                test_file.read(100)  # 读取文件头部验证文件可访问性
        except IOError as e:
            return False, f"无法打开文件: {str(e)}"

        # 读取Excel文件（只读取第一个工作表）
        df = pd.read_excel(file_path, sheet_name=0, dtype=str)
        if df.empty:
            return False, "Excel文件为空或未包含数据"

        # 特殊处理：对奖惩信息和家庭成员信息表处理合并单元格
        if table_name in ['rewards', 'family']:
            logger.info(f"开始处理 {table_name_mapping[table_name]} 表的合并单元格...")

            if file_ext == '.xlsx':
                # 处理.xlsx格式文件
                try:
                    # 使用openpyxl获取合并单元格信息
                    wb = load_workbook(file_path, data_only=True)  # 添加data_only=True
                    sheet = wb.active
                    merged_ranges = list(sheet.merged_cells.ranges)  # 转换为列表避免修改时的异常

                    logger.info(f"找到 {len(merged_ranges)} 个合并单元格")

                    # 填充合并单元格的值
                    for merged_range in merged_ranges:
                        min_row, min_col, max_row, max_col = (
                            merged_range.min_row,
                            merged_range.min_col,
                            merged_range.max_row,
                            merged_range.max_col
                        )

                        # 跳过标题行（第一行）
                        if min_row == 1:
                            continue

                        # 获取合并区域的左上角单元格值
                        value = sheet.cell(row=min_row, column=min_col).value

                        # 填充到合并区域的所有单元格
                        for r in range(min_row, max_row + 1):
                            for c in range(min_col, max_col + 1):
                                # 转换Excel索引到DataFrame索引
                                df_row = r - 2  # 行索引：标题行占1行，所以减2
                                df_col = c - 1  # 列索引：0-based

                                if df_row >= 0 and df_row < len(df) and df_col < len(df.columns):
                                    # 直接赋值，不再检查列名
                                    df.iat[df_row, df_col] = str(value) if value else ""

                    logger.info(f"成功处理 {len(merged_ranges)} 个合并单元格（.xlsx格式）")
                except Exception as merge_error:
                    logger.error(f"处理.xlsx合并单元格时出错: {str(merge_error)}", exc_info=True)
                    logger.warning("将继续导入，但合并单元格可能未被正确处理")

            elif file_ext == '.xls':
                # 处理.xls格式文件
                try:
                    # 使用xlrd打开文件
                    book = xlrd.open_workbook(file_path)
                    sheet = book.sheet_by_index(0)

                    # 获取合并单元格信息
                    merged_cells = sheet.merged_cells

                    logger.info(f"找到 {len(merged_cells)} 个合并单元格")

                    # 填充合并单元格的值
                    for (min_row, max_row, min_col, max_col) in merged_cells:
                        # 跳过标题行（第一行）
                        if min_row == 0:
                            continue

                        # 获取合并区域的左上角单元格值
                        value = sheet.cell_value(min_row, min_col)

                        # 填充到合并区域的所有单元格
                        for r in range(min_row, max_row):
                            for c in range(min_col, max_col):
                                # 转换索引（xlrd的行列索引都是0-based）
                                if r >= 1:  # 跳过标题行
                                    df_row = r - 1  # 标题行占1行，所以减1
                                    df_col = c
                                    if df_row < len(df) and df_col < len(df.columns):
                                        # 直接赋值，不再检查列名
                                        df.iat[df_row, df_col] = str(value) if value else ""

                    logger.info(f"成功处理 {len(merged_cells)} 个合并单元格（.xls格式）")
                except Exception as merge_error:
                    logger.error(f"处理.xls合并单元格时出错: {str(merge_error)}", exc_info=True)
                    logger.warning("将继续导入，但合并单元格可能未被正确处理")


        # 清理列名
        df.columns = [clean_column_name(c) for c in df.columns]

        # 记录处理后的列名
        logger.info(f"处理后的列名: {list(df.columns)}")

        # 删除全空行
        df = df.dropna(how='all')
        if df.empty:
            return False, "删除空行后数据为空"

        # 记录导入后的数据样本
        logger.debug(f"导入后数据样本: \n{df.head(5).to_string()}")

        # ==== 新增：处理base_info表的年度考核字段 ====
        year_to_index = {}  # 年份到通用标记的映射

        if table_name == 'base_info':
            # 1. 识别年度考核字段
            assessment_years = []

            for col in df.columns:
                match = re.search(r'(\d{4})年年度考核结果', col)
                if match:
                    year = int(match.group(1))
                    assessment_years.append(year)

            # 2. 验证是否为连续五年
            if assessment_years:
                assessment_years.sort()
                if len(assessment_years) != 5:
                    return False, "必须包含连续的五个年度考核字段"

                for i in range(1, 5):
                    if assessment_years[i] - assessment_years[i - 1] != 1:
                        return False, "年度考核字段必须为连续五年"

                # 3. 检查年份配置是否已存在
                existing_years = db.get_assessment_years()
                if existing_years and existing_years != assessment_years:
                    return False, f"年度考核区间不匹配（已配置: {existing_years}，当前: {assessment_years}），请先清空数据库"

                # 4. 存储年份配置
                if not existing_years:
                    if not db.set_assessment_years(assessment_years):
                        return False, "保存年度考核配置失败"

                # 5. 创建年份到通用标记的映射
                year_to_index = {year: f"assessment_{idx}" for idx, year in enumerate(assessment_years)}

        # 转换为字典列表
        records: List[Dict[str, Any]] = []
        for _, row in df.iterrows():
            record = {}
            for col_name, value in row.items():
                # 清理列名
                cleaned_name = clean_column_name(col_name)

                # 年度考核字段特殊处理
                match = re.search(r'(\d{4})年年度考核结果', col_name)
                if match and table_name == 'base_info':
                    year = int(match.group(1))
                    if year in year_to_index:
                        record[year_to_index[year]] = convert_excel_date(value)
                        continue  # 跳过常规处理

                # 常规字段处理
                record[cleaned_name] = convert_excel_date(value)

            records.append(record)

        # 导入到指定表
        db.import_excel_data(table_name, records)
        count = len(records)
        return True, f"成功导入{table_name_mapping[table_name]} {count} 条记录"

    except Exception as e:
        logger.error(f"导入{table_name}失败: {e}", exc_info=True)
        return False, f"导入{table_name_mapping[table_name]}失败: {e}"